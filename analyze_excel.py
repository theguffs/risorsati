import sys
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def analyze_excel(file_path):
    """Analizza il file Excel e restituisce informazioni sulla struttura"""
    try:
        print(f"\nCaricamento file: {file_path}...")
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        
        print(f"File caricato!\n")
        print(f"Numero di fogli: {len(wb.sheetnames)}\n")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n{'='*70}")
            print(f"FOGLIO: {sheet_name}")
            print(f"{'='*70}")
            
            # Analisi struttura (solo prime righe)
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            print(f"Dimensioni: {max_row} righe x {max_col} colonne")
            
            # Leggi solo le prime righe per capire la struttura
            headers = []
            if max_row > 0 and max_col > 0:
                print(f"\nIntestazioni colonne:")
                header_row = []
                for col in range(1, min(max_col + 1, 50)):  # Limita a 50 colonne
                    cell = sheet.cell(row=1, column=col)
                    value = cell.value
                    header_row.append(value if value else "")
                    if value:
                        print(f"  [{col:2d}] {value}")
                
                headers = header_row
            
            # Leggi alcune righe di esempio (solo prime 5 righe con dati)
            print(f"\nDati di esempio (prime 5 righe):")
            row_count = 0
            for row_idx in range(2, min(max_row + 1, 20)):  # Analizza prime 20 righe
                row_data = {}
                has_data = False
                
                for col in range(1, min(max_col + 1, 50)):
                    cell = sheet.cell(row=row_idx, column=col)
                    value = cell.value
                    if value is not None and str(value).strip():
                        has_data = True
                        header_name = headers[col-1] if col-1 < len(headers) and headers[col-1] else f"Col{col}"
                        row_data[header_name] = value
                
                if has_data and row_count < 5:
                    print(f"\n  -- Riga {row_idx} --")
                    for key, val in list(row_data.items())[:15]:  # Mostra max 15 campi
                        print(f"    {key}: {val}")
                    if len(row_data) > 15:
                        print(f"    ... e altri {len(row_data) - 15} campi")
                    row_count += 1
            
            # Cerca formule (solo prime righe)
            print(f"\nAnalisi formule (prime 100 righe)...")
            formula_count = 0
            formula_examples = []
            for row in sheet.iter_rows(min_row=1, max_row=min(100, max_row), values_only=False):
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        formula_count += 1
                        if len(formula_examples) < 10:
                            header = headers[cell.column - 1] if cell.column - 1 < len(headers) else f"Col{cell.column}"
                            formula_examples.append({
                                "cell": cell.coordinate,
                                "header": header,
                                "formula": cell.value
                            })
            
            if formula_count > 0:
                print(f"  Trovate {formula_count} formule (prime 100 righe)")
                print(f"\n  Esempi formule:")
                for ex in formula_examples[:5]:
                    print(f"    {ex['cell']} ({ex['header']}): {ex['formula']}")
            else:
                print(f"  Nessuna formula trovata (prime 100 righe)")
        
        wb.close()
        print(f"\n{'='*70}")
        print("Analisi completata!")
        print(f"{'='*70}\n")
        
    except Exception as e:
        print(f"Errore durante la lettura: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    file_path = "File Extra&Fissi 2025 - NEW.xlsx"
    analyze_excel(file_path)
